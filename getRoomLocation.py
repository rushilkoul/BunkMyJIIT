import os
import sys
import json
from openpyxl import load_workbook

EXCEL_FILE = "RoomLocation.xlsx"
JSON_FILE = "room_lookup.json"

room_lookup = {}

def generate_json_from_excel():
    """Generate room_lookup.json from Excel and return the dict"""
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: {EXCEL_FILE} not found.")
        return {}

    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        room_id = str(row[headers["ROOMID"] - 1]).strip()
        building = row[headers["BUILDING"] - 1]
        floor = row[headers["FLOOR"] - 1]
        if room_id:  # skip empty room IDs
            lookup[room_id] = f"{building} ({floor})"

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(lookup, f, ensure_ascii=False, indent=2)

    print(f"Generated {JSON_FILE} from {EXCEL_FILE} ({len(lookup)} rooms)")
    return lookup


def load_room_lookup():
    """Load room_lookup.json into memory"""
    global room_lookup
    if os.path.exists(JSON_FILE):
        with open(JSON_FILE, "r", encoding="utf-8") as f:
            room_lookup = {k.strip(): v for k, v in json.load(f).items()}
        print(f"Loaded {JSON_FILE} ({len(room_lookup)} rooms)")
    else:
        print(f"{JSON_FILE} not found. Run with -generate to create it.")
        room_lookup = {}


def getLocation(roomID):
    """Return location string for a roomID, or None if not found"""
    global room_lookup
    if not room_lookup:
        load_room_lookup()
    return room_lookup.get(str(roomID).strip(), None)


# CLI support
if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "-generate":
        room_lookup = generate_json_from_excel()
    else:
        load_room_lookup()
        print("Room lookup loaded. Example usage:")
        print(">>> getLocation('FF5') ->", getLocation("FF5"))
